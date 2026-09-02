from datetime import date, datetime, timedelta, time
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.timezone import now

from openpyxl import Workbook
from openpyxl.styles import Font

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from backend.members.decorators import admin_required
from backend.members.models import (
    ClaimSettlement,
    Member,
    Payment,
    PaymentRequest,
)
from backend.members.views_admin.admin_finance import get_report_dates

# ======================================================
# TREASURER REPORT DATA
# ======================================================

def get_treasurer_report_data(request):
    """
    Build the Treasurer Control Panel report data.

    This function is shared by the PDF and Excel exports.

    Financial and compliance calculations must match
    the Treasurer Control Panel.
    """

    # ==================================================
    # REPORTING PERIOD
    # ==================================================

    (
        start_date,
        end_date,
        period,
    ) = get_report_dates(request)

    period_labels = {
        "today": "Today",
        "this_month": "This Month",
        "last_month": "Last Month",
        "this_year": "Current Financial Year",
        "last_year": "Previous Financial Year",
        "custom": "Custom Range",
    }

    period_label = period_labels.get(
        period,
        "This Month"
    )

    # ==================================================
    # COMPLETED PAYMENTS
    #
    # Same logic as Treasurer Control Panel.
    # ==================================================

    payments_queryset = Payment.objects.filter(
        status=Payment.STATUS_COMPLETED,
        paid_at__date__range=[
            start_date,
            end_date,
        ],
    )

    # ==================================================
    # TOTAL FUNDS COLLECTED
    #
    # Same source and calculation as the
    # Treasurer Control Panel.
    # ==================================================

    total_collected = (
        payments_queryset.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    # ==================================================
    # COMPLIANCE
    #
    # Same calculation as Treasurer Control Panel.
    #
    # Members who made completed payments during
    # the selected reporting period.
    # ==================================================

    active_members = Member.objects.filter(
        status=Member.STATUS_ACTIVE
    ).count()

    members_who_paid = (
        payments_queryset
        .exclude(member__isnull=True)
        .values("member")
        .distinct()
        .count()
    )

    compliance_rate = 0

    if active_members > 0:

        compliance_rate = int(
            (
                members_who_paid
                / active_members
            ) * 100
        )

    # ==================================================
    # APPROVED CLAIM SETTLEMENTS
    #
    # Only approved claim settlements represent
    # actual financial payouts.
    #
    # Same logic as Treasurer Control Panel.
    # ==================================================

    claims_paid_queryset = ClaimSettlement.objects.filter(
        is_approved=True,
        settlement_date__date__range=[
            start_date,
            end_date,
        ],
    )

    # ==================================================
    # CLAIM SETTLEMENT TOTALS
    # ==================================================

    claims_paid_total = Decimal("0.00")

    deductions_total = Decimal("0.00")

    collected_total = Decimal("0.00")

    for settlement in claims_paid_queryset:

        claims_paid_total += (
            settlement.amount_paid
            or Decimal("0.00")
        )

        deductions_total += (
            settlement.total_deductions
            or Decimal("0.00")
        )

        collected_total += (
            settlement.total_collected
            or Decimal("0.00")
        )

    # ==================================================
    # RETURN REPORT DATA
    # ==================================================

    return {

        # ==================================================
        # REPORTING PERIOD
        # ==================================================

        "period": period,

        "period_label": period_label,

        "start_date": start_date,

        "end_date": end_date,

        # ==================================================
        # FINANCIAL SUMMARY
        # ==================================================

        "total_collected": total_collected,

        "compliance_rate": compliance_rate,

        # ==================================================
        # CLAIM SETTLEMENTS
        # ==================================================

        "claims_paid_total": claims_paid_total,

        "deductions_total": deductions_total,

        "collected_total": collected_total,
    }

# ======================================================
# TREASURER PDF EXPORT
# ======================================================

@admin_required
def treasurer_dashboard_pdf(request):

    # ==================================================
    # GET REPORT DATA
    # ==================================================

    context = get_treasurer_report_data(request)

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4
    )

    styles = getSampleStyleSheet()

    elements = []

    # ==================================================
    # TITLE
    # ==================================================

    title = Paragraph(
        """
        <font size="20">
        <b>Treasurer Financial Report</b>
        </font>
        """,
        styles["Title"],
    )

    elements.append(title)

    elements.append(Spacer(1, 20))

    # ==================================================
    # REPORTING PERIOD
    # ==================================================

    report_period = Paragraph(
        f"""
        <b>Reporting Period:</b>
        {context["start_date"].strftime("%d/%m/%Y")}
        to
        {context["end_date"].strftime("%d/%m/%Y")}
        """,
        styles["Normal"],
    )

    elements.append(report_period)

    elements.append(Spacer(1, 20))

    # ==================================================
    # TABLE DATA
    # ==================================================

    data = [

        ["Metric", "Value"],

        [
            "Funds Collected",
            f"£{context['collected_total']:,.2f}"
        ],

        [
            "Claims Paid",
            f"£{context['claims_paid_total']:,.2f}"
        ],

        [
            "Total Deductions",
            f"£{context['deductions_total']:,.2f}"
        ],

        [
            "Compliance",
            f"{context['compliance_rate']}%"
        ],
    ]

    # ==================================================
    # TABLE
    # ==================================================

    table = Table(
        data,
        colWidths=[300, 180]
    )

    table.setStyle(TableStyle([

        ("BACKGROUND",
         (0, 0),
         (-1, 0),
         colors.HexColor("#1f2937")),

        ("TEXTCOLOR",
         (0, 0),
         (-1, 0),
         colors.white),

        ("GRID",
         (0, 0),
         (-1, -1),
         1,
         colors.grey),

        ("FONTNAME",
         (0, 0),
         (-1, 0),
         "Helvetica-Bold"),

        ("BOTTOMPADDING",
         (0, 0),
         (-1, 0),
         12),

    ]))

    elements.append(table)

    # ==================================================
    # BUILD PDF
    # ==================================================

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    # ==================================================
    # RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="treasurer_report.pdf"'
    )

    response.write(pdf)

    return response

# ======================================================
# TREASURER EXCEL EXPORT
# ======================================================

@admin_required
def treasurer_dashboard_excel(request):

    # ==================================================
    # GET REPORT DATA
    # ==================================================

    context = get_treasurer_report_data(request)

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Treasurer Report"

    # ==================================================
    # TITLE
    # ==================================================

    sheet["A1"] = "Treasurer Financial Report"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    # ==================================================
    # REPORT PERIOD
    # ==================================================

    sheet["A3"] = "Reporting Period"

    sheet["B3"] = (
        f"{context['start_date'].strftime('%d/%m/%Y')} "
        f"to "
        f"{context['end_date'].strftime('%d/%m/%Y')}"
    )

    # ==================================================
    # HEADERS
    # ==================================================

    sheet["A5"] = "Metric"

    sheet["B5"] = "Value"

    sheet["A5"].font = Font(bold=True)

    sheet["B5"].font = Font(bold=True)

    # ==================================================
    # DATA
    # ==================================================

    rows = [

        [
            "Funds Collected",
            float(context["collected_total"])
        ],

        [
            "Claims Paid",
            float(context["claims_paid_total"])
        ],

        [
            "Total Deductions",
            float(context["deductions_total"])
        ],

        [
            "Compliance",
            context["compliance_rate"]
        ],
    ]

    row_number = 6

    for row in rows:

        sheet.cell(
            row=row_number,
            column=1,
            value=row[0]
        )

        sheet.cell(
            row=row_number,
            column=2,
            value=row[1]
        )

        row_number += 1

    # ==================================================
    # COLUMN WIDTHS
    # ==================================================

    sheet.column_dimensions["A"].width = 35

    sheet.column_dimensions["B"].width = 25

    # ==================================================
    # RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="treasurer_report.xlsx"'
    )

    workbook.save(response)

    return response

# ======================================================
# TREASURER DASHBOARD
# ======================================================

@admin_required
def treasurer_dashboard(request):

    today = now().date()

    # ==================================================
    # DEFAULT PERIOD:
    # CURRENT FINANCIAL YEAR
    #
    # Financial Year:
    # 1 June -> 31 May
    # ==================================================

    period = request.GET.get(
        "period",
        "this_year"
    )

    # ==================================================
    # THIS MONTH
    # ==================================================

    if period == "this_month":

        start_date = today.replace(day=1)

        end_date = today

    # ==================================================
    # LAST MONTH
    # ==================================================

    elif period == "last_month":

        first_this_month = today.replace(day=1)

        end_date = (
            first_this_month - timedelta(days=1)
        )

        start_date = end_date.replace(day=1)

    # ==================================================
    # LAST 90 DAYS
    # ==================================================

    elif period == "90_days":

        start_date = today - timedelta(days=90)

        end_date = today

    # ==================================================
    # LAST 6 MONTHS
    # ==================================================

    elif period == "6_months":

        start_date = today - timedelta(days=180)

        end_date = today

    # ==================================================
    # CURRENT FINANCIAL YEAR
    # ==================================================
    # Financial Year:
    # 1 June -> 31 May
    # ==================================================

    elif period == "this_year":

        # ----------------------------------------------
        # Example:
        #
        # Today = 24 May 2026
        #
        # Current FY:
        # 01/06/2025 -> 24/05/2026
        # ----------------------------------------------

        if today.month >= 6:

            fy_start = today.year

        else:

            fy_start = today.year - 1

        start_date = date(
            fy_start,
            6,
            1
        )

        end_date = today

    # ==================================================
    # PREVIOUS FINANCIAL YEAR
    # ==================================================

    elif period == "last_year":

        # ----------------------------------------------
        # Example:
        #
        # Today = 24 May 2026
        #
        # Previous FY:
        # 01/06/2024 -> 31/05/2025
        # ----------------------------------------------

        if today.month >= 6:

            previous_fy_start = today.year - 1

        else:

            previous_fy_start = today.year - 2

        start_date = date(
            previous_fy_start,
            6,
            1
        )

        end_date = date(
            previous_fy_start + 1,
            5,
            31
        )

    # ==================================================
    # CUSTOM RANGE
    # ==================================================

    elif period == "custom":

        start_date = request.GET.get(
            "start_date"
        )

        end_date = request.GET.get(
            "end_date"
        )

    # ==================================================
    # DEFAULT FALLBACK
    # ==================================================

    else:

        start_date = today.replace(day=1)

        end_date = today

    # ==================================================
    # PAYMENTS
    #
    # Incoming funds only
    # ==================================================

    payments = Payment.objects.filter(
        paid_at__date__range=[
            start_date,
            end_date
        ]
    )

    # ==================================================
    # TOTAL FUNDS COLLECTED
    #
    # Excludes claim payouts
    # ==================================================

    collected_total = payments.filter(
        payment_type__in=[
            "membership",
            "subscription",
            "other",
        ]
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ==================================================
    # CLAIM SETTLEMENTS
    # ==================================================

    claim_settlements = ClaimSettlement.objects.filter(
        settlement_date__range=[
            start_date,
            end_date
        ]
    )

    # ==================================================
    # TOTAL DEDUCTIONS
    # ==================================================

    deductions_total = sum(
        settlement.total_deductions
        for settlement in claim_settlements
    )

    # ==================================================
    # NET CLAIMS PAID
    #
    # amount_paid is a PROPERTY
    # not a database field.
    # Must be calculated in Python.
    # ==================================================

    claims_paid_total = sum(
        settlement.amount_paid
        for settlement in claim_settlements
    )

    # ==================================================
    # COMPLIANCE
    # ==================================================
    # Members who made membership/subscription
    # payments during the selected period.
    # ==================================================

    active_members = Member.objects.filter(
        is_active=True
    )

    total_members = active_members.count()

    compliant_members = active_members.filter(
        payment__payment_type__in=[
            "membership",
            "subscription",
        ],
        payment__paid_at__date__range=[
            start_date,
            end_date
        ]
    ).distinct().count()

    compliance_rate = 0

    if total_members > 0:

        compliance_rate = round(
            (compliant_members / total_members) * 100,
            1
        )

    # ==================================================
    # CONTEXT
    # ==================================================

    context = {

        "period": period,

        "start_date": start_date,

        "end_date": end_date,

        "collected_total": collected_total,

        "deductions_total": deductions_total,

        "claims_paid_total": claims_paid_total,

        "compliance_rate": compliance_rate,
    }

    return render(
        request,
        "members/admin/finance/admin_treasurer_dashboard.html",
        context
    )