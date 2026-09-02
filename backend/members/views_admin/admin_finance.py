from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Sum, Count
from .admin_auth import admin_required
from django.utils import timezone
from datetime import datetime
from backend.members.models import Payment, Claim, Member, PaymentRequest, ClaimSettlement
from django.db.models.functions import TruncMonth
from backend.members.views_admin.admin_auth import admin_required  
from datetime import timedelta, datetime, time, date
from decimal import Decimal
import json
from io import BytesIO
from django.utils.timezone import now
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font
from calendar import monthrange
from django.utils.dateparse import parse_date

@admin_required
def treasurer_control_panel(request):
    """
    Treasurer financial dashboard.
    """

    # ======================================================
    # REPORTING PERIOD
    # ======================================================

    (
        start_date,
        end_date,
        period,
    ) = get_report_dates(request)

    period_labels = {
        "today": "Today",
        "this_month": "This Month",
        "last_month": "Last Month",
        "this_year": "Current Financial Year",
        "last_year": "Previous Financial Year",
        "custom": "Custom Range",
    }

    period_label = period_labels.get(
        period,
        "This Month"
    )

    # ======================================================
    # COMPLETED PAYMENTS — SELECTED PERIOD
    # ======================================================

    payments_queryset = Payment.objects.filter(
        status=Payment.STATUS_COMPLETED,
        paid_at__date__range=[
            start_date,
            end_date,
        ],
    )

    # ======================================================
    # TOTAL FUNDS COLLECTED
    # ======================================================

    total_collected = (
        payments_queryset.aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
    )

    # ======================================================
    # OUTSTANDING LIABILITIES
    # ======================================================

    outstanding_requests = PaymentRequest.objects.filter(
        claim__status="approved"
    )

    outstanding_liabilities = sum(
        (
            pr.amount or Decimal("0.00")
        )
        for pr in outstanding_requests
    )

    # ======================================================
    # COMPLIANCE RATE — SELECTED PERIOD
    # ======================================================

    active_members = Member.objects.filter(
        status=Member.STATUS_ACTIVE
    ).count()

    members_who_paid = (
        payments_queryset
        .exclude(member__isnull=True)
        .values("member")
        .distinct()
        .count()
    )

    compliance_rate = 0

    if active_members > 0:

        compliance_rate = int(
            (
                members_who_paid
                / active_members
            ) * 100
        )

    # ======================================================
    # APPROVED CLAIM SETTLEMENTS — SELECTED PERIOD
    #
    # Only approved claim settlements represent
    # actual financial payouts.
    # ======================================================

    claims_paid_queryset = ClaimSettlement.objects.filter(
        is_approved=True,
        settlement_date__date__range=[
            start_date,
            end_date,
        ],
    )

    # ======================================================
    # CLAIM SETTLEMENT TOTALS
    # ======================================================

    claims_paid_total = Decimal("0.00")
    deductions_total = Decimal("0.00")
    collected_total = Decimal("0.00")

    for settlement in claims_paid_queryset:

        claims_paid_total += (
            settlement.amount_paid
            or Decimal("0.00")
        )

        deductions_total += (
            settlement.total_deductions
            or Decimal("0.00")
        )

        collected_total += (
            settlement.total_collected
            or Decimal("0.00")
        )

    # ======================================================
    # NET POSITION
    # ======================================================

    net_position = (
        total_collected
        - claims_paid_total
    )

    # ======================================================
    # CONTEXT
    # ======================================================

    context = {

        # ==================================================
        # REPORTING PERIOD
        # ==================================================

        "period": period,

        "period_label": period_label,

        "report_start": start_date,

        "report_end": end_date,

        # Preserve submitted custom dates
        # for the date input fields.

        "start_date": request.GET.get(
            "start_date",
            ""
        ),

        "end_date": request.GET.get(
            "end_date",
            ""
        ),

        # ==================================================
        # FINANCIAL SUMMARY
        # ==================================================

        "total_collected": total_collected,

        "outstanding_liabilities": outstanding_liabilities,

        "net_position": net_position,

        "compliance_rate": compliance_rate,

        # ==================================================
        # CLAIM SETTLEMENTS
        # ==================================================

        "claims_paid_total": claims_paid_total,

        "deductions_total": deductions_total,

        "collected_total": collected_total,
    }

    return render(
        request,
        "members/admin/finance/admin_treasurer_dashboard.html",
        context,
    )

@admin_required
def treasurer_analytics_dashboard(request):
    """
    Treasurer analytics dashboard.

    Produces period-filtered:
    - Monthly income trend
    - Claims paid trend
    - Payment request trend

    All datasets are serialized safely for Chart.js.
    """

    # ======================================================
    # REPORTING PERIOD
    # ======================================================

    (
        start_date,
        end_date,
        period,
    ) = get_report_dates(request)

    period_labels = {
        "today": "Today",
        "this_month": "This Month",
        "last_month": "Last Month",
        "this_year": "Current Financial Year",
        "last_year": "Previous Financial Year",
        "custom": "Custom Range",
    }

    period_label = period_labels.get(
        period,
        "This Month"
    )

    # ======================================================
    # MONTHLY INCOME TREND
    # ======================================================
    # Source of truth:
    # Completed Payment records using paid_at.
    # ======================================================

    income_queryset = (
        Payment.objects
        .filter(
            status=Payment.STATUS_COMPLETED,
            paid_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("paid_at")
        )
        .values("month")
        .annotate(
            total=Sum("amount")
        )
        .order_by("month")
    )

    income_trend = []

    for item in income_queryset:

        income_trend.append({

            "month": (
                item["month"].strftime("%Y-%m-%d")
                if item["month"]
                else ""
            ),

            "total": float(
                item["total"] or 0
            ),
        })

    # ======================================================
    # CLAIMS PAID TREND
    # ======================================================
    # Source of truth:
    # Approved ClaimSettlement records using settlement_date.
    #
    # Only approved settlements represent actual
    # financial payouts.
    # ======================================================

    claims_queryset = (
        ClaimSettlement.objects
        .filter(
            is_approved=True,
            settlement_date__isnull=False,
            settlement_date__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("settlement_date")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    claims_trend = []

    for item in claims_queryset:

        claims_trend.append({

            "month": (
                item["month"].strftime("%Y-%m-%d")
                if item["month"]
                else ""
            ),

            "total": item["total"],
        })

    # ======================================================
    # PAYMENT REQUEST TREND
    # ======================================================
    # Source of truth:
    # PaymentRequest.created_at.
    # ======================================================

    requests_queryset = (
        PaymentRequest.objects
        .filter(
            created_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("created_at")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    request_trend = []

    for item in requests_queryset:

        request_trend.append({

            "month": (
                item["month"].strftime("%Y-%m-%d")
                if item["month"]
                else ""
            ),

            "total": item["total"],
        })

    # ======================================================
    # CONTEXT
    # ======================================================

    context = {

        # ==================================================
        # REPORTING PERIOD
        # ==================================================

        "period": period,

        "period_label": period_label,

        "report_start": start_date,

        "report_end": end_date,

        # Preserve submitted custom dates
        # for the filter inputs.

        "start_date": request.GET.get(
            "start_date",
            ""
        ),

        "end_date": request.GET.get(
            "end_date",
            ""
        ),

        # ==================================================
        # JSON SAFE ANALYTICS DATA
        # ==================================================

        "income_trend": json.dumps(
            income_trend
        ),

        "claims_trend": json.dumps(
            claims_trend
        ),

        "request_trend": json.dumps(
            request_trend
        ),
    }

    return render(
        request,
        "members/admin/finance/admin_treasurer_analytics.html",
        context,
    )

# ======================================================
# HELPER:
# DETERMINE REPORT PERIOD
# ======================================================

def get_report_dates(request):

    today = now().date()

    period = request.GET.get(
        "period",
        "this_month"
    )

    # TODAY
    if period == "today":

        start_date = today
        end_date = today

    # THIS MONTH
    elif period == "this_month":

        start_date = today.replace(day=1)
        end_date = today

    # LAST MONTH
    elif period == "last_month":

        first_this_month = today.replace(day=1)

        last_month_end = (
            first_this_month - timedelta(days=1)
        )

        start_date = last_month_end.replace(day=1)

        end_date = last_month_end

    # THIS YEAR
    elif period == "this_year":

        # ----------------------------------------------
        # Financial Year:
        # 1 June -> 31 May
        # ----------------------------------------------

        if today.month >= 6:

            # Current FY started this year

            start_date = date(
                today.year,
                6,
                1
            )

            end_date = today

        else:

            # Current FY started last year

            start_date = date(
                today.year - 1,
                6,
                1
            )

            end_date = today

    # LAST YEAR
    elif period == "last_year":
        # ----------------------------------------------
        # Previous Financial Year
        # ----------------------------------------------
        if today.month >= 6:

            # Previous FY:
            # 1 Jun last year -> 31 May this year

            start_date = date(
                today.year - 1,
                6,
                1
            )

            end_date = date(
                today.year,
                5,
                31
            )

        else:

            # Previous FY:
            # 1 Jun two years ago -> 31 May last year

            start_date = date(
                today.year - 2,
                6,
                1
            )

            end_date = date(
                today.year - 1,
                5,
                31
            )

    # CUSTOM RANGE
    elif period == "custom":

        start_date = parse_date(
            request.GET.get("start_date", "")
        )

        end_date = parse_date(
            request.GET.get("end_date", "")
        )

    # DEFAULT
    else:

        start_date = today.replace(day=1)
        end_date = today

    return (
        start_date,
        end_date,
        period
    )

# ======================================================
# TREASURER ANALYTICS PDF EXPORT
# ======================================================

@admin_required
def treasurer_analytics_pdf(request):
    """
    Export Treasurer Financial Analytics as PDF.

    Uses the same reporting-period logic and source data
    as treasurer_analytics_dashboard.
    """

    # ==================================================
    # REPORTING PERIOD
    # ==================================================

    (
        start_date,
        end_date,
        period,
    ) = get_report_dates(request)

    period_labels = {
        "today": "Today",
        "this_month": "This Month",
        "last_month": "Last Month",
        "this_year": "Current Financial Year",
        "last_year": "Previous Financial Year",
        "custom": "Custom Range",
    }

    period_label = period_labels.get(
        period,
        "This Month"
    )

    # ==================================================
    # MONTHLY INCOME
    # ==================================================

    income_queryset = (
        Payment.objects
        .filter(
            status=Payment.STATUS_COMPLETED,
            paid_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("paid_at")
        )
        .values("month")
        .annotate(
            total=Sum("amount")
        )
        .order_by("month")
    )

    income_data = []

    for item in income_queryset:

        income_data.append([

            item["month"].strftime(
                "%B %Y"
            ) if item["month"] else "",

            f"£{item['total'] or Decimal('0.00'):.2f}",
        ])

    # ==================================================
    # CLAIMS PAID
    # ==================================================

    claims_queryset = (
        ClaimSettlement.objects
        .filter(
            is_approved=True,
            settlement_date__isnull=False,
            settlement_date__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("settlement_date")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    claims_data = []

    for item in claims_queryset:

        claims_data.append([

            item["month"].strftime(
                "%B %Y"
            ) if item["month"] else "",

            item["total"],
        ])

    # ==================================================
    # PAYMENT REQUESTS
    # ==================================================

    requests_queryset = (
        PaymentRequest.objects
        .filter(
            created_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("created_at")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    requests_data = []

    for item in requests_queryset:

        requests_data.append([

            item["month"].strftime(
                "%B %Y"
            ) if item["month"] else "",

            item["total"],
        ])

    # ==================================================
    # PDF RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="treasurer_financial_analytics.pdf"'
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()

    story = []

    # ==================================================
    # TITLE
    # ==================================================

    story.append(
        Paragraph(
            "Treasurer Financial Analytics",
            styles["Title"],
        )
    )

    story.append(
        Spacer(1, 12)
    )

    story.append(
        Paragraph(
            f"<b>Reporting Period:</b> "
            f"{period_label}",
            styles["Normal"],
        )
    )

    story.append(
        Paragraph(
            f"<b>Date Range:</b> "
            f"{start_date.strftime('%d/%m/%Y')} "
            f"- "
            f"{end_date.strftime('%d/%m/%Y')}",
            styles["Normal"],
        )
    )

    story.append(
        Spacer(1, 20)
    )

    # ==================================================
    # MONTHLY INCOME TABLE
    # ==================================================

    story.append(
        Paragraph(
            "Monthly Income",
            styles["Heading2"],
        )
    )

    story.append(
        Spacer(1, 8)
    )

    income_table_data = [

        ["Month", "Income"],

    ] + (
        income_data
        if income_data
        else [["No data", "£0.00"]]
    )

    income_table = Table(
        income_table_data,
        colWidths=[250, 150],
    )

    income_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "ALIGN",
                (1, 1),
                (1, -1),
                "RIGHT",
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
        ])
    )

    story.append(
        income_table
    )

    story.append(
        Spacer(1, 24)
    )

    # ==================================================
    # CLAIMS PAID TABLE
    # ==================================================

    story.append(
        Paragraph(
            "Claims Paid Per Month",
            styles["Heading2"],
        )
    )

    story.append(
        Spacer(1, 8)
    )

    claims_table_data = [

        ["Month", "Claims Paid"],

    ] + (
        claims_data
        if claims_data
        else [["No data", 0]]
    )

    claims_table = Table(
        claims_table_data,
        colWidths=[250, 150],
    )

    claims_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "ALIGN",
                (1, 1),
                (1, -1),
                "RIGHT",
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
        ])
    )

    story.append(
        claims_table
    )

    story.append(
        Spacer(1, 24)
    )

    # ==================================================
    # PAYMENT REQUESTS TABLE
    # ==================================================

    story.append(
        Paragraph(
            "Payment Requests Created",
            styles["Heading2"],
        )
    )

    story.append(
        Spacer(1, 8)
    )

    requests_table_data = [

        ["Month", "Requests Created"],

    ] + (
        requests_data
        if requests_data
        else [["No data", 0]]
    )

    requests_table = Table(
        requests_table_data,
        colWidths=[250, 150],
    )

    requests_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "ALIGN",
                (1, 1),
                (1, -1),
                "RIGHT",
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8,
            ),
        ])
    )

    story.append(
        requests_table
    )

    # ==================================================
    # BUILD PDF
    # ==================================================

    document.build(
        story
    )

    pdf = buffer.getvalue()

    buffer.close()

    response.write(
        pdf
    )

    return response

# ======================================================
# TREASURER ANALYTICS EXCEL EXPORT
# ======================================================

@admin_required
def treasurer_analytics_excel(request):
    """
    Export Treasurer Financial Analytics as Excel.

    Uses the same reporting-period logic and source data
    as treasurer_analytics_dashboard.
    """

    # ==================================================
    # REPORTING PERIOD
    # ==================================================

    (
        start_date,
        end_date,
        period,
    ) = get_report_dates(request)

    period_labels = {
        "today": "Today",
        "this_month": "This Month",
        "last_month": "Last Month",
        "this_year": "Current Financial Year",
        "last_year": "Previous Financial Year",
        "custom": "Custom Range",
    }

    period_label = period_labels.get(
        period,
        "This Month"
    )

    # ==================================================
    # MONTHLY INCOME
    # ==================================================

    income_queryset = (
        Payment.objects
        .filter(
            status=Payment.STATUS_COMPLETED,
            paid_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("paid_at")
        )
        .values("month")
        .annotate(
            total=Sum("amount")
        )
        .order_by("month")
    )

    # ==================================================
    # CLAIMS PAID
    # ==================================================

    claims_queryset = (
        ClaimSettlement.objects
        .filter(
            is_approved=True,
            settlement_date__isnull=False,
            settlement_date__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("settlement_date")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    # ==================================================
    # PAYMENT REQUESTS
    # ==================================================

    requests_queryset = (
        PaymentRequest.objects
        .filter(
            created_at__date__range=[
                start_date,
                end_date,
            ],
        )
        .annotate(
            month=TruncMonth("created_at")
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    # ==================================================
    # WORKBOOK
    # ==================================================

    workbook = Workbook()

    # Remove default worksheet.
    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    # ==================================================
    # MONTHLY INCOME SHEET
    # ==================================================

    income_sheet = workbook.create_sheet(
        "Monthly Income"
    )

    income_sheet.append([
        "Treasurer Financial Analytics"
    ])

    income_sheet.append([
        f"Reporting Period: {period_label}"
    ])

    income_sheet.append([
        (
            "Date Range: "
            f"{start_date.strftime('%d/%m/%Y')} "
            f"- "
            f"{end_date.strftime('%d/%m/%Y')}"
        )
    ])

    income_sheet.append([])

    income_sheet.append([
        "Month",
        "Income (£)",
    ])

    for item in income_queryset:

        month_value = (
            item["month"].strftime("%B %Y")
            if item["month"]
            else ""
        )

        income_sheet.append([

            month_value,

            item["total"]
            or Decimal("0.00"),
        ])

    # Header formatting

    for cell in income_sheet[5]:

        cell.font = Font(
            bold=True
        )

    income_sheet.column_dimensions[
        "A"
    ].width = 25

    income_sheet.column_dimensions[
        "B"
    ].width = 18

    # Currency formatting

    for row in range(
        6,
        income_sheet.max_row + 1
    ):

        income_sheet[
            f"B{row}"
        ].number_format = "£#,##0.00"

    # ==================================================
    # CLAIMS PAID SHEET
    # ==================================================

    claims_sheet = workbook.create_sheet(
        "Claims Paid"
    )

    claims_sheet.append([
        "Treasurer Financial Analytics"
    ])

    claims_sheet.append([
        f"Reporting Period: {period_label}"
    ])

    claims_sheet.append([
        (
            "Date Range: "
            f"{start_date.strftime('%d/%m/%Y')} "
            f"- "
            f"{end_date.strftime('%d/%m/%Y')}"
        )
    ])

    claims_sheet.append([])

    claims_sheet.append([
        "Month",
        "Claims Paid",
    ])

    for item in claims_queryset:

        month_value = (
            item["month"].strftime("%B %Y")
            if item["month"]
            else ""
        )

        claims_sheet.append([

            month_value,

            item["total"],
        ])

    for cell in claims_sheet[5]:

        cell.font = Font(
            bold=True
        )

    claims_sheet.column_dimensions[
        "A"
    ].width = 25

    claims_sheet.column_dimensions[
        "B"
    ].width = 18

    # ==================================================
    # PAYMENT REQUESTS SHEET
    # ==================================================

    requests_sheet = workbook.create_sheet(
        "Payment Requests"
    )

    requests_sheet.append([
        "Treasurer Financial Analytics"
    ])

    requests_sheet.append([
        f"Reporting Period: {period_label}"
    ])

    requests_sheet.append([
        (
            "Date Range: "
            f"{start_date.strftime('%d/%m/%Y')} "
            f"- "
            f"{end_date.strftime('%d/%m/%Y')}"
        )
    ])

    requests_sheet.append([])

    requests_sheet.append([
        "Month",
        "Requests Created",
    ])

    for item in requests_queryset:

        month_value = (
            item["month"].strftime("%B %Y")
            if item["month"]
            else ""
        )

        requests_sheet.append([

            month_value,

            item["total"],
        ])

    for cell in requests_sheet[5]:

        cell.font = Font(
            bold=True
        )

    requests_sheet.column_dimensions[
        "A"
    ].width = 25

    requests_sheet.column_dimensions[
        "B"
    ].width = 20

    # ==================================================
    # HTTP RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="treasurer_financial_analytics.xlsx"'
    )

    workbook.save(
        response
    )

    return response

# ======================================================
# MAIN FINANCE SUMMARY
# ======================================================

@admin_required
def finance_summary(request):

    (
        start_date,
        end_date,
        selected_period
    ) = get_report_dates(request)

    # ==================================================
    # FILTERED PAYMENTS
    # ==================================================

    payments = Payment.objects.filter(
        status=Payment.STATUS_COMPLETED,
        paid_at__date__range=[
            start_date,
            end_date
        ]
    )

    # ==================================================
    # TOTALS
    # ==================================================

    total_received = payments.aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_claims = payments.filter(
        payment_type="claim"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_membership = payments.filter(
        payment_type="membership"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_subscription = payments.filter(
        payment_type="subscription"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_other = payments.filter(
        payment_type="other"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ==================================================
    # OUTSTANDING REQUESTS
    # ==================================================

    outstanding_requests = PaymentRequest.objects.filter(
        status="active"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ==================================================
    # TOTAL PAID OUT
    # ==================================================

    claim_settlements = ClaimSettlement.objects.filter(
        is_approved=True,
        settlement_date__gte=start_date,
        settlement_date__lte=end_date,
    )

    total_paid_out = sum(
        settlement.amount_paid
        for settlement in claim_settlements
    )

    net_position = (
        total_received
        - total_paid_out
    )

    completed_payments = payments.count()

    # ==================================================
    # EXPORTS
    # ==================================================

    export_type = request.GET.get("export")

    if export_type == "pdf":

        return finance_summary_pdf(
        request,
        start_date,
        end_date,
        total_received,
        total_claims,
        total_membership,
        total_subscription,
        total_other,
        outstanding_requests,
        total_paid_out,
        completed_payments,
    )

    if export_type == "excel":

        return finance_summary_excel(
            request,
            start_date,
            end_date,
            total_received,
            total_claims,
            total_membership,
            total_subscription,
            total_other,
            outstanding_requests,
            total_paid_out,
            completed_payments,
        )

    # ==================================================
    # QUERY PARAMS
    # ==================================================

    query_params = request.GET.copy()

    if "export" in query_params:
        query_params.pop("export")

    context = {

        "report_start": start_date,
        "report_end": end_date,

        "selected_period": selected_period,

        "query_params": query_params.urlencode(),

        "total_received": total_received,

        "total_claims": total_claims,

        "total_membership": total_membership,

        "total_subscription": total_subscription,

        "total_other": total_other,

        "outstanding_requests": outstanding_requests,

        "total_paid_out": total_paid_out,

        "net_position": net_position,

        "completed_payments": completed_payments,
    }

    return render(
        request,
        "members/admin/admin_finance_summary.html",
        context
    )

# ======================================================
# PDF EXPORT
# ======================================================

@admin_required
def finance_summary_pdf(
    request,
    start_date,
    end_date,
    total_received,
    total_claims,
    total_membership,
    total_subscription,
    total_other,
    outstanding_requests,
    total_paid_out,
    completed_payments,
):

    # ==================================================
    # CREATE PDF BUFFER
    # ==================================================

    buffer = BytesIO()

    # ==================================================
    # PDF DOCUMENT SETUP
    # ==================================================

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    # ==================================================
    # REPORT TITLE
    # ==================================================

    title = Paragraph(
        """
        <font size="20">
        <b>KRO Welfare Management</b>
        </font>
        <br/>
        Financial Summary Report
        """,
        styles["Title"],
    )

    elements.append(title)

    elements.append(Spacer(1, 20))

    # ==================================================
    # REPORTING PERIOD
    # ==================================================

    report_period = Paragraph(
        f"""
        <b>Reporting Period:</b>
        {start_date.strftime("%d/%m/%Y")}
        to
        {end_date.strftime("%d/%m/%Y")}
        """,
        styles["Normal"],
    )

    elements.append(report_period)

    elements.append(Spacer(1, 20))

    # ==================================================
    # TABLE DATA
    # ==================================================

    data = [

        ["Metric", "Value"],

        ["Total Received",
         f"£{total_received:,.2f}"],

        ["Total Received - Claims",
         f"£{total_claims:,.2f}"],

        ["Total Received - Membership",
         f"£{total_membership:,.2f}"],

        ["Total Received - Subscription",
         f"£{total_subscription:,.2f}"],

        ["Total Received - Other",
         f"£{total_other:,.2f}"],

        ["Outstanding Requests",
         f"£{outstanding_requests:,.2f}"],

        ["Total Paid Out",
         f"£{total_paid_out:,.2f}"],

        ["Payments Completed",
         f"{completed_payments:,}"],
    ]

    # ==================================================
    # CREATE TABLE
    # ==================================================

    table = Table(
        data,
        colWidths=[320, 160]
    )

    # ==================================================
    # TABLE STYLING
    # ==================================================

    table.setStyle(
        TableStyle([

            # HEADER
            ("BACKGROUND",
             (0, 0),
             (-1, 0),
             colors.HexColor("#1f2937")),

            ("TEXTCOLOR",
             (0, 0),
             (-1, 0),
             colors.white),

            ("FONTNAME",
             (0, 0),
             (-1, 0),
             "Helvetica-Bold"),

            ("FONTSIZE",
             (0, 0),
             (-1, 0),
             12),

            ("BOTTOMPADDING",
             (0, 0),
             (-1, 0),
             12),

            # BODY
            ("BACKGROUND",
             (0, 1),
             (-1, -1),
             colors.whitesmoke),

            ("FONTNAME",
             (0, 1),
             (-1, -1),
             "Helvetica"),

            ("FONTSIZE",
             (0, 1),
             (-1, -1),
             11),

            ("GRID",
             (0, 0),
             (-1, -1),
             1,
             colors.grey),

            ("ALIGN",
             (1, 1),
             (-1, -1),
             "RIGHT"),

            ("BOTTOMPADDING",
             (0, 1),
             (-1, -1),
             8),
        ])
    )

    elements.append(table)

    elements.append(Spacer(1, 30))

    # ==================================================
    # FOOTER
    # ==================================================

    footer = Paragraph(
        """
        Generated automatically by
        KRO Welfare Management System
        """,
        styles["Italic"],
    )

    elements.append(footer)

    # ==================================================
    # BUILD PDF
    # ==================================================

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    # ==================================================
    # HTTP RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="finance_summary.pdf"'
    )

    response.write(pdf)

    return response

# ======================================================
# EXCEL EXPORT
# ======================================================

@admin_required
def finance_summary_excel(
    request,
    start_date,
    end_date,
    total_received,
    total_claims,
    total_membership,
    total_subscription,
    total_other,
    outstanding_requests,
    total_paid_out,
    completed_payments,
):

    # ==================================================
    # CREATE WORKBOOK
    # ==================================================

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Finance Summary"

    # ==================================================
    # REPORT TITLE
    # ==================================================

    sheet["A1"] = "KRO Welfare Management"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet["A3"] = "Financial Summary Report"

    sheet["A3"].font = Font(
        bold=True,
        size=13
    )

    # ==================================================
    # REPORT PERIOD
    # ==================================================

    sheet["A5"] = "Reporting Period"

    sheet["A5"].font = Font(bold=True)

    sheet["B5"] = (
        f"{start_date.strftime('%d/%m/%Y')} "
        f"to "
        f"{end_date.strftime('%d/%m/%Y')}"
    )

    # ==================================================
    # TABLE HEADERS
    # ==================================================

    sheet["A7"] = "Metric"

    sheet["B7"] = "Value"

    sheet["A7"].font = Font(bold=True)
    sheet["B7"].font = Font(bold=True)

    # ==================================================
    # TABLE DATA
    # ==================================================

    rows = [

        ["Total Received",
         float(total_received)],

        ["Total Received - Claims",
         float(total_claims)],

        ["Total Received - Membership",
         float(total_membership)],

        ["Total Received - Subscription",
         float(total_subscription)],

        ["Total Received - Other",
         float(total_other)],

        ["Outstanding Requests",
         float(outstanding_requests)],

        ["Total Paid Out",
         float(total_paid_out)],

        ["Payments Completed",
         completed_payments],
    ]

    # ==================================================
    # WRITE DATA TO SHEET
    # ==================================================

    row_number = 8

    for row in rows:

        sheet.cell(
            row=row_number,
            column=1,
            value=row[0]
        )

        sheet.cell(
            row=row_number,
            column=2,
            value=row[1]
        )

        row_number += 1

    # ==================================================
    # COLUMN WIDTHS
    # ==================================================

    sheet.column_dimensions["A"].width = 40

    sheet.column_dimensions["B"].width = 25

    # ==================================================
    # RESPONSE
    # ==================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="finance_summary.xlsx"'
    )

    workbook.save(response)

    return response
